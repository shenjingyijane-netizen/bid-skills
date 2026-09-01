#!/usr/bin/env python3
import argparse
import json
import os
import tempfile
import zipfile
from pathlib import Path

OPENPYXL_IMPORT_ERROR = None
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.comments import Comment
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.page import PageMargins
except ModuleNotFoundError as exc:
    OPENPYXL_IMPORT_ERROR = exc


FONT = "Microsoft YaHei"
WHITE = "FFFFFF"
NAVY = "243447"
GREEN = "2F6B4F"
RED = "9C0006"
AMBER = "9C6500"
PALE_BLUE = "EDF3F8"
PALE_GREEN = "E2F0D9"
PALE_RED = "FCE8E6"
PALE_YELLOW = "FFF2CC"
GRAY = "F3F4F6"
if OPENPYXL_IMPORT_ERROR is None:
    THIN = Side(style="thin", color="B8C2CC")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
else:
    THIN = None
    BORDER = None
CONTRACT_RESULTS = {
    "一致",
    "明确冲突",
    "疑似冲突",
    "合同未覆盖",
    "项目需求未覆盖",
    "待确认",
}
CONTRACT_STATUSES = {"待确认", "已确认有冲突", "已确认无冲突"}
PACKAGE_ROOT = Path(__file__).resolve().parent.parent
REQUIRED_PACKAGE_ASSETS = (
    "SKILL.md",
    "agents/openai.yaml",
    "assets/项目介绍_通用空白模板.xlsx",
    "references/pre-bid-schema.md",
    "references/renderer-input-schema.md",
    "scripts/build_project_snapshot.py",
)


def text(value):
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(text(item) for item in value)
    return str(value)


def records(data, key):
    value = data.get(key, [])
    if not isinstance(value, list) or any(not isinstance(item, dict) for item in value):
        raise ValueError(f"{key} must be a list of objects")
    return value


def preflight_package():
    missing = []
    for relative_path in REQUIRED_PACKAGE_ASSETS:
        asset = PACKAGE_ROOT / relative_path
        if not asset.is_file() or asset.stat().st_size == 0:
            missing.append(relative_path)
    if missing:
        raise FileNotFoundError(
            "incomplete tender-project-snapshot package; missing assets: "
            + ", ".join(missing)
        )
    template = PACKAGE_ROOT / "assets/项目介绍_通用空白模板.xlsx"
    if not zipfile.is_zipfile(template):
        raise ValueError("the bundled Excel visual template is not a valid xlsx file")
    if OPENPYXL_IMPORT_ERROR is not None:
        raise RuntimeError(
            "Python dependency openpyxl is unavailable; this environment cannot generate xlsx"
        ) from OPENPYXL_IMPORT_ERROR
    return True


def merged_row(
    ws,
    row,
    end_col,
    value,
    fill,
    font_color="444444",
    size=9,
    bold=False,
    height=26,
    horizontal="center",
):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    cell = ws.cell(row, 1, value)
    cell.font = Font(name=FONT, size=size, bold=bold, color=font_color)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(
        horizontal=horizontal,
        vertical="center",
        wrap_text=True,
        indent=1 if horizontal == "left" else 0,
    )
    ws.row_dimensions[row].height = height


def section(ws, row, value, fill):
    merged_row(ws, row, 10, value, fill, WHITE, 11, True, 24)


def header(ws, row, labels, fill):
    for col, label in enumerate(labels, 1):
        cell = ws.cell(row, col, label)
        cell.font = Font(name=FONT, size=9, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 34


def style_row(ws, row, cols=10, fill=WHITE, height=40):
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.font = Font(name=FONT, size=9, color="222222")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = height


def spacer(ws, row):
    ws.row_dimensions[row].height = 8


def page_setup(ws, last_row, repeat_rows):
    ws.sheet_view.showGridLines = False
    ws.print_area = f"A1:J{last_row}"
    ws.print_title_rows = repeat_rows
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.25, right=0.25, top=0.35, bottom=0.35, header=0.15, footer=0.15
    )


def validate_contract_comparison(data):
    contract = data.get("contract_comparison")
    if not isinstance(contract, dict):
        raise ValueError("contract_comparison must be an object")
    present = contract.get("contract_present")
    completed = contract.get("comparison_completed")
    items = contract.get("items")
    if not isinstance(present, bool) or not isinstance(completed, bool):
        raise ValueError("contract_present and comparison_completed must be booleans")
    if not isinstance(items, list) or any(not isinstance(item, dict) for item in items):
        raise ValueError("contract_comparison.items must be a list of objects")
    if not present:
        if completed or items:
            raise ValueError("a missing contract cannot have completed comparison items")
        if not text(contract.get("contract_source")).strip():
            raise ValueError("contract_source must explain that no contract was found")
        return contract
    if not completed:
        raise ValueError("contract comparison is required when a contract is present")
    if not text(contract.get("contract_source")).strip():
        raise ValueError("contract_source is required when a contract is present")
    if not items:
        raise ValueError("contract comparison items are required when a contract is present")
    if text(contract.get("template_note")).strip() and not text(
        contract.get("template_note_source")
    ).strip():
        raise ValueError("template_note_source is required when template_note is provided")
    required = (
        "subject",
        "requirement_text",
        "contract_text",
        "requirement_source",
        "contract_source",
    )
    for index, item in enumerate(items, 1):
        missing = [key for key in required if not text(item.get(key)).strip()]
        if missing:
            raise ValueError(
                f"contract comparison item {index} is missing: {', '.join(missing)}"
            )
        if item.get("result") not in CONTRACT_RESULTS:
            raise ValueError(f"invalid contract comparison result at item {index}")
        status = item.get("manual_status", "待确认")
        if status not in CONTRACT_STATUSES:
            raise ValueError(f"invalid manual_status at contract item {index}")
    return contract


def validate_initial_delivery_state(data, contract):
    qualifications = records(data, "qualifications")
    for index, item in enumerate(qualifications, 1):
        if text(item.get("user_choice")).strip():
            raise ValueError(
                f"qualification {index} has a prefilled user_choice; formal delivery must start blank"
            )
    for index, item in enumerate(contract.get("items", []), 1):
        if item.get("manual_status", "待确认") != "待确认":
            raise ValueError(
                f"contract item {index} has a confirmed manual_status; formal delivery must start 待确认"
            )


def make_intro(wb, data, contract):
    qualifications = records(data, "qualifications")
    requirements = records(data, "project_requirements")
    quote_rules = records(data, "quote_restrictions")
    contract_items = contract["items"]
    contract_issues = [item for item in contract_items if item.get("result") != "一致"]
    overview = data.get("overview", {})
    if not isinstance(overview, dict):
        raise ValueError("overview must be an object")

    ws = wb.active
    ws.title = "项目介绍"
    ws.sheet_view.showGridLines = False
    widths = [6, 15, 28, 14, 18, 12, 14, 17, 18, 20]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.merge_cells("A1:J1")
    ws["A1"] = "项目介绍与投标准入"
    ws["A1"].font = Font(name=FONT, size=16, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 31
    merged_row(
        ws,
        2,
        10,
        "通读完整文件后生成；资格区只放企业级投标准入条件，报名购标和服务资源不得混入。",
        PALE_BLUE,
        height=30,
        horizontal="left",
    )

    ws.merge_cells("A3:C3")
    ws["A3"] = "投标资格二次审查总结果"
    ws["A3"].font = Font(name=FONT, size=11, bold=True, color=WHITE)
    ws["A3"].fill = PatternFill("solid", fgColor=RED)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("D3:J3")
    if qualifications:
        first_qual = 7
        last_qual = first_qual + len(qualifications) - 1
        ws["D3"] = (
            f'=IF(COUNTIFS(B{first_qual}:B{last_qual},"<>",F{first_qual}:F{last_qual},"否")>0,'
            '"停止：存在不符合项",'
            f'IF(COUNTIFS(B{first_qual}:B{last_qual},"<>",F{first_qual}:F{last_qual},"")>0,'
            '"暂停：存在待确认项","通过：可以继续项目分析"))'
        )
    else:
        ws["D3"] = "未发现企业级投标准入条件"
    ws["D3"].font = Font(name=FONT, size=11, bold=True, color="222222")
    ws["D3"].fill = PatternFill("solid", fgColor=PALE_YELLOW)
    ws["D3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["D3"].comment = Comment(
        "任一已生成资格问题选择“否”即停止；存在未选择项则暂停。", "Codex"
    )
    ws.row_dimensions[3].height = 28
    ws.conditional_formatting.add(
        "D3:J3", FormulaRule(formula=['LEFT($D$3,2)="通过"'], fill=PatternFill("solid", fgColor=PALE_GREEN))
    )
    ws.conditional_formatting.add(
        "D3:J3", FormulaRule(formula=['LEFT($D$3,2)="停止"'], fill=PatternFill("solid", fgColor=PALE_RED))
    )
    ws.conditional_formatting.add(
        "D3:J3",
        FormulaRule(
            formula=['OR(LEFT($D$3,2)="暂停",LEFT($D$3,2)="未发")'],
            fill=PatternFill("solid", fgColor=PALE_YELLOW),
        ),
    )

    section(ws, 4, "A. 企业级投标准入审查", RED)
    merged_row(
        ws,
        5,
        10,
        "每行只核对一个企业事实；选择“是”代表通过。获取文件、人员设备及服务承诺不得作为资格停止项。",
        PALE_RED,
        height=32,
        horizontal="left",
    )
    qual_headers = [
        "序号",
        "具体核验问题",
        "",
        "核对材料/查询入口",
        "",
        "用户选择",
        "自动状态",
        "不符合后果",
        "原文位置",
        "备注",
    ]
    header(ws, 6, qual_headers, RED)
    ws.merge_cells("B6:C6")
    ws.merge_cells("D6:E6")
    ws["B6"] = "具体核验问题"
    ws["D6"] = "核对材料/查询入口"

    row = 7
    qual_start = row
    for index, item in enumerate(qualifications, 1):
        style_row(ws, row, fill=GRAY if index % 2 == 0 else WHITE, height=40)
        ws.cell(row, 1, index)
        ws.cell(row, 2, text(item.get("question")))
        ws.cell(row, 4, text(item.get("check_material")))
        ws.cell(row, 6, "")
        ws.cell(row, 7, f'=IF(F{row}="","待确认",IF(F{row}="是","通过","不通过"))')
        ws.cell(row, 8, text(item.get("consequence")))
        ws.cell(row, 9, text(item.get("source")))
        ws.cell(row, 10, text(item.get("notes")))
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        for col in (1, 6, 7):
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row += 1
    qual_end = row - 1
    if qualifications:
        dv = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
        dv.error = "请选择“是”或“否”。"
        dv.errorTitle = "无效选择"
        dv.prompt = "每行只回答一个问题；选择“是”代表通过。"
        dv.promptTitle = "企业级投标准入核验"
        dv.showErrorMessage = True
        dv.showInputMessage = True
        ws.add_data_validation(dv)
        dv.add(f"F{qual_start}:F{qual_end}")
        ws.conditional_formatting.add(
            f"A{qual_start}:J{qual_end}",
            FormulaRule(formula=[f'$F{qual_start}="是"'], fill=PatternFill("solid", fgColor=PALE_GREEN)),
        )
        ws.conditional_formatting.add(
            f"A{qual_start}:J{qual_end}",
            FormulaRule(formula=[f'$F{qual_start}="否"'], fill=PatternFill("solid", fgColor=PALE_RED)),
        )
        ws.conditional_formatting.add(
            f"A{qual_start}:J{qual_end}",
            FormulaRule(
                formula=[f'AND($B{qual_start}<>"",$F{qual_start}="")'],
                fill=PatternFill("solid", fgColor=PALE_YELLOW),
            ),
        )

    spacer(ws, row)
    row += 1
    overview_section = row
    section(ws, row, "B. 项目30秒概览", NAVY)
    row += 1
    overview_header = row
    overview_headers = [
        "项目名称",
        "项目编号",
        "项目一句话定义",
        "采购人",
        "招标代理机构",
        "采购方式",
        "中标/成交数量",
        "投标截止时间",
        "开标地点",
        "原文位置",
    ]
    header(ws, row, overview_headers, NAVY)
    row += 1
    overview_data = row
    style_row(ws, row, fill=PALE_BLUE, height=52)
    overview_keys = [
        "project_name",
        "project_number",
        "one_sentence_summary",
        "purchaser",
        "procurement_agent",
        "procurement_method",
        "award_count",
        "bid_deadline",
        "opening_location",
        "source",
    ]
    for col, key in enumerate(overview_keys, 1):
        ws.cell(row, col, text(overview.get(key)) or "未说明")
    row += 1

    spacer(ws, row)
    row += 1
    requirement_section = row
    section(ws, row, "C. 项目内容与关键要求", GREEN)
    row += 1
    requirement_header = row
    requirement_headers = [
        "序号",
        "项目组成",
        "采购范围/关键要求",
        "服务/交付对象",
        "时间",
        "地点",
        "数量/规模",
        "标准/成果",
        "验收方式",
        "原文位置",
    ]
    header(ws, row, requirement_headers, GREEN)
    row += 1
    shown_requirements = requirements or [{"component": "项目内容", "requirement": "未说明"}]
    requirement_start = row
    req_keys = [
        "component",
        "requirement",
        "recipient",
        "time",
        "location",
        "scale",
        "standard_or_deliverable",
        "acceptance",
        "source",
    ]
    for index, item in enumerate(shown_requirements, 1):
        style_row(ws, row, fill=PALE_GREEN if index % 2 == 1 else WHITE, height=44)
        ws.cell(row, 1, index)
        for col, key in enumerate(req_keys, 2):
            ws.cell(row, col, text(item.get(key)) or "未说明")
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
        row += 1
    requirement_end = row - 1

    spacer(ws, row)
    row += 1
    contract_section = row
    section(ws, row, "D. 合同核对", AMBER)
    row += 1
    contract_summary = row
    if not contract["contract_present"]:
        contract_status = "不适用"
        summary = "文件未提供合同模板，不适用。"
        summary_fill = GRAY
    elif contract_issues:
        contract_status = "待人工确认"
        summary = (
            f"已完成 {len(contract_items)} 项合同与项目需求比对，"
            f"发现 {len(contract_issues)} 项需要人工确认。"
        )
        summary_fill = PALE_YELLOW
    else:
        contract_status = "已完成"
        summary = (
            f"已完成 {len(contract_items)} 项合同与项目需求比对，"
            "未发现需要人工确认的差异。"
        )
        summary_fill = PALE_GREEN
    template_note = (
        "不适用"
        if not contract["contract_present"]
        else text(contract.get("template_note")) or "未说明"
    )
    source_parts = [text(contract.get("contract_source"))]
    note_source = text(contract.get("template_note_source"))
    if note_source and note_source not in source_parts:
        source_parts.append(note_source)
    summary_text = (
        f"核对状态：{contract_status}\n"
        f"核对结论：{summary}\n"
        f"合同模板适用说明：{template_note}\n"
        f"原文位置：{'；'.join(part for part in source_parts if part)}"
    )
    merged_row(
        ws,
        row,
        10,
        summary_text,
        summary_fill,
        size=9,
        height=58,
        horizontal="left",
    )
    row += 1

    contract_detail_header = None
    contract_detail_start = row
    contract_detail_end = row - 1
    if contract_issues:
        contract_detail_header = row
        contract_headers = [
            "序号",
            "核对事项",
            "项目需求原文",
            "",
            "合同原文",
            "",
            "核对结果",
            "待确认问题/可能影响",
            "原文位置",
            "人工确认",
        ]
        header(ws, row, contract_headers, AMBER)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws.cell(row, 3, "项目需求原文")
        ws.cell(row, 5, "合同原文")
        row += 1
        contract_detail_start = row
        for index, item in enumerate(contract_issues, 1):
            fill = PALE_RED if item.get("result") == "明确冲突" else PALE_YELLOW
            style_row(ws, row, fill=fill, height=54)
            ws.cell(row, 1, index)
            ws.cell(row, 2, text(item.get("subject")))
            ws.cell(row, 3, text(item.get("requirement_text")))
            ws.cell(row, 5, text(item.get("contract_text")))
            ws.cell(row, 7, text(item.get("result")))
            detail = "\n".join(
                value
                for value in [
                    text(item.get("confirmation_question")),
                    text(item.get("impact")),
                ]
                if value
            )
            ws.cell(row, 8, detail or "待人工确认")
            ws.cell(
                row,
                9,
                f"需求：{text(item.get('requirement_source'))}\n"
                f"合同：{text(item.get('contract_source'))}",
            )
            ws.cell(row, 10, "待确认")
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
            for col in (1, 7, 10):
                ws.cell(row, col).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            row += 1
        contract_detail_end = row - 1
        dv = DataValidation(
            type="list",
            formula1='"待确认,已确认有冲突,已确认无冲突"',
            allow_blank=False,
        )
        dv.error = "请选择列表中的确认状态。"
        dv.errorTitle = "无效选择"
        dv.showErrorMessage = True
        ws.add_data_validation(dv)
        dv.add(f"J{contract_detail_start}:J{contract_detail_end}")
        ws.conditional_formatting.add(
            f"A{contract_detail_start}:J{contract_detail_end}",
            FormulaRule(
                formula=[f'$J{contract_detail_start}="已确认有冲突"'],
                fill=PatternFill("solid", fgColor=PALE_RED),
            ),
        )
        ws.conditional_formatting.add(
            f"A{contract_detail_start}:J{contract_detail_end}",
            FormulaRule(
                formula=[f'$J{contract_detail_start}="已确认无冲突"'],
                fill=PatternFill("solid", fgColor=PALE_GREEN),
            ),
        )

    spacer(ws, row)
    row += 1
    quote_section = row
    section(ws, row, "E. 报价限制提示", NAVY)
    row += 1
    merged_row(
        ws,
        row,
        10,
        "只提取影响报价合规性的限制；实际报价使用原文件规定的报价载体，不重建报价明细。",
        PALE_BLUE,
        height=30,
        horizontal="left",
    )
    row += 1
    quote_header = row
    quote_headers = [
        "序号",
        "适用范围",
        "限制类型",
        "文件明确要求",
        "",
        "含税要求",
        "原始报价载体",
        "",
        "不符合后果",
        "原文位置",
    ]
    header(ws, row, quote_headers, NAVY)
    ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
    ws.cell(row, 4, "文件明确要求")
    ws.cell(row, 7, "原始报价载体")
    row += 1
    shown_quotes = quote_rules or [{"scope": "项目", "restriction_type": "未说明"}]
    quote_start = row
    for index, item in enumerate(shown_quotes, 1):
        style_row(ws, row, fill=PALE_BLUE if index % 2 == 1 else WHITE, height=42)
        ws.cell(row, 1, index)
        ws.cell(row, 2, text(item.get("scope")) or "未说明")
        ws.cell(row, 3, text(item.get("restriction_type")) or "未说明")
        ws.cell(row, 4, text(item.get("requirement")) or "未说明")
        ws.cell(row, 6, text(item.get("tax_requirement")) or "未说明")
        ws.cell(row, 7, text(item.get("quote_carrier")) or "未说明")
        ws.cell(row, 9, text(item.get("consequence")) or "未说明")
        ws.cell(row, 10, text(item.get("source")) or "未说明")
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
        ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
        row += 1
    quote_end = row - 1

    title_rows = [
        1,
        3,
        4,
        6,
        overview_section,
        overview_header,
        requirement_section,
        requirement_header,
        contract_section,
        quote_section,
        quote_header,
    ]
    if contract_detail_header is not None:
        title_rows.append(contract_detail_header)
    for title_row in title_rows:
        for cell in ws[title_row]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A7"
    page_setup(ws, quote_end, "1:2")
    return {
        "qual_start": qual_start,
        "qual_end": qual_end,
        "overview_header": overview_header,
        "overview_data": overview_data,
        "requirement_header": requirement_header,
        "requirement_start": requirement_start,
        "requirement_end": requirement_end,
        "contract_section": contract_section,
        "contract_summary": contract_summary,
        "contract_detail_header": contract_detail_header,
        "contract_detail_start": contract_detail_start,
        "contract_detail_end": contract_detail_end,
        "contract_issue_count": len(contract_issues),
        "quote_header": quote_header,
        "quote_start": quote_start,
        "quote_end": quote_end,
    }


def validate_workbook(path, layout):
    with zipfile.ZipFile(path) as archive:
        if archive.testzip() is not None:
            raise ValueError("generated xlsx archive is corrupt")
    wb = load_workbook(path, data_only=False)
    expected = ["项目介绍"]
    if wb.sheetnames != expected:
        raise ValueError(f"unexpected sheets: {wb.sheetnames}")
    ws = wb["项目介绍"]
    checks = {
        "A1": "项目介绍与投标准入",
        "A6": "序号",
        f"A{layout['overview_header']}": "项目名称",
        f"A{layout['requirement_header']}": "序号",
        f"A{layout['contract_section']}": "D. 合同核对",
        f"A{layout['quote_header']}": "序号",
    }
    for coordinate, expected_value in checks.items():
        if ws[coordinate].value != expected_value:
            raise ValueError(f"layout validation failed at {coordinate}")
    if ws.max_column != 10:
        raise ValueError(f"unexpected column count: {ws.max_column}")
    if ws["A1"].alignment.horizontal != "center":
        raise ValueError("main title is not centered")
    if "核对状态：" not in text(ws.cell(layout["contract_summary"], 1).value):
        raise ValueError("contract comparison summary is missing")
    for rules in ws.conditional_formatting._cf_rules.values():
        for rule in rules:
            dxf = wb._differential_styles.styles[rule.dxfId] if rule.dxfId is not None else rule.dxf
            if dxf is not None and dxf.font is not None:
                raise ValueError("conditional formatting must not change font color")
    if layout["qual_end"] >= layout["qual_start"]:
        validations = [str(item.sqref) for item in ws.data_validations.dataValidation]
        expected_range = (
            f"F{layout['qual_start']}"
            if layout["qual_start"] == layout["qual_end"]
            else f"F{layout['qual_start']}:F{layout['qual_end']}"
        )
        if expected_range not in validations:
            raise ValueError("qualification dropdown range is misaligned")
        for row in range(layout["qual_start"], layout["qual_end"] + 1):
            if ws.cell(row, 6).value not in (None, ""):
                raise ValueError(f"qualification choice is prefilled at F{row}")
            if not str(ws.cell(row, 7).value).startswith("=IF(F"):
                raise ValueError(f"qualification formula is misaligned at G{row}")
    if layout["contract_issue_count"]:
        if ws.cell(layout["contract_detail_header"], 1).value != "序号":
            raise ValueError("contract comparison detail header is misaligned")
        validations = [str(item.sqref) for item in ws.data_validations.dataValidation]
        expected_range = (
            f"J{layout['contract_detail_start']}"
            if layout["contract_detail_start"] == layout["contract_detail_end"]
            else f"J{layout['contract_detail_start']}:J{layout['contract_detail_end']}"
        )
        if expected_range not in validations:
            raise ValueError("contract confirmation dropdown range is misaligned")
        for row in range(layout["contract_detail_start"], layout["contract_detail_end"] + 1):
            if ws.cell(row, 10).value != "待确认":
                raise ValueError(f"contract confirmation is not reset at J{row}")
    return True


def build_workbook(data, output_path):
    preflight_package()
    if not isinstance(data, dict):
        raise ValueError("input JSON root must be an object")
    contract = validate_contract_comparison(data)
    validate_initial_delivery_state(data, contract)
    wb = Workbook()
    layout = make_intro(wb, data, contract)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    output = Path(output_path).expanduser().resolve()
    if output.suffix.lower() != ".xlsx":
        raise ValueError("output path must end with .xlsx")
    output.parent.mkdir(parents=True, exist_ok=True)
    handle = tempfile.NamedTemporaryFile(
        prefix=f".{output.stem}-", suffix=".xlsx", dir=output.parent, delete=False
    )
    temp_path = Path(handle.name)
    handle.close()
    try:
        wb.save(temp_path)
        validate_workbook(temp_path, layout)
        os.replace(temp_path, output)
    finally:
        if temp_path.exists():
            temp_path.unlink()
    return output


def main():
    parser = argparse.ArgumentParser(description="Build a stable tender project snapshot workbook")
    parser.add_argument("--preflight", action="store_true", help="verify bundled Skill assets")
    parser.add_argument("input_json", nargs="?", help="structured extraction JSON")
    parser.add_argument("output_xlsx", nargs="?", help="output .xlsx path")
    args = parser.parse_args()
    if args.preflight:
        preflight_package()
        print("tender-project-snapshot package preflight passed")
        return
    if not args.input_json or not args.output_xlsx:
        parser.error("input_json and output_xlsx are required unless --preflight is used")
    with open(args.input_json, "r", encoding="utf-8") as source:
        data = json.load(source)
    print(build_workbook(data, args.output_xlsx))


if __name__ == "__main__":
    try:
        main()
    except (FileNotFoundError, RuntimeError, ValueError) as exc:
        raise SystemExit(f"ERROR: {exc}") from None
